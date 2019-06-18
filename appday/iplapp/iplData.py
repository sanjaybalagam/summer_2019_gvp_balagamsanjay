import openpyxl
import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'appday.settings')
django.setup()
from iplapp.models import *

import datetime

def date_modifier(date):
    if '-' in date:
        field1,field2,field3=date.split('-')
        if len(field1) == 4:
            return date
        return field3+'-'+field2+'-'+field1
    field1, field2, field3 = date.split('/')
    return '20'+field3+'-'+field2+'-'+field1

def main():
    # wb = openpyxl.load_workbook('matches.xlsx')
    # match_sheet = wb.active
    # for i in range(2, match_sheet.max_row + 1):
    #     if match_sheet.cell(row=i,column=1).value == None:
    #         continue
    #     m=Matches()
    #
    #     m.matchid=match_sheet.cell(row=i,column=1).value
    #     if m.matchid is "":
    #         continue
    #     m.season = match_sheet.cell(row=i, column=2).value
    #     m.city = match_sheet.cell(row=i, column=3).value
    #     m.date =match_sheet.cell(row=i, column=4).value
    #     m.team1 = match_sheet.cell(row=i,column=5).value
    #     m.team2 = match_sheet.cell(row=i,column=6).value
    #     m.toss_winner = match_sheet.cell(row=i,column=7).value
    #     m.toss_decision = match_sheet.cell(row=i,column=8).value
    #     m.result = match_sheet.cell(row=i,column=9).value
    #     m.dl_applied = match_sheet.cell(row=i,column=10).value
    #     m.winner = match_sheet.cell(row=i,column=11).value
    #     m.win_by_runs= match_sheet.cell(row=i,column=12).value
    #     m.win_by_wickets = match_sheet.cell(row=i,column=13).value
    #     m.player_of_match = match_sheet.cell(row=i,column=14).value
    #     m.venue = match_sheet.cell(row=i,column=15).value
    #     m.umpire1 = match_sheet.cell(row=i,column=16).value
    #     m.umpire2 = match_sheet.cell(row=i,column=17).value
    #     m.umpire3 = match_sheet.cell(row=i,column=18).value
    #
    #     m.save()
    # wb.close()
    import MySQLdb
    db = MySQLdb.connect(host="127.0.0.1", user="root", passwd="admin", database="hackday")
    wb = openpyxl.load_workbook('deliveries.xlsx')
    detail_sheet=wb.active
    cur = db.cursor()
    print('starting')
    for i in range(2,179080):
        # d=Balls()
        # d.inning = detail_sheet.cell(row=i,column=2).value
        # d.batting_team = detail_sheet.cell(row=i,column=3).value
        # d.bowling_team = detail_sheet.cell(row=i,column=4).value
        # d.over = detail_sheet.cell(row=i,column=5).value
        # d.ball = detail_sheet.cell(row=i,column=6).value
        # d.batsman = detail_sheet.cell(row=i,column=7).value
        # d.non_striker = detail_sheet.cell(row=i,column=8).value
         bowler = detail_sheet.cell(row=i,column=9).value
         cur.execute('update iplapp_balls set bowler="'+bowler+'" where id='+ str(i-1) +';')

        # d.is_super_over = detail_sheet.cell(row=i,column=10).value
        # d.wide_runs = detail_sheet.cell(row=i,column=11).value
        # d.bye_runs = detail_sheet.cell(row=i,column=12).value
        # d.legbye_runs = detail_sheet.cell(row=i,column=13).value
        # d.noball_runs = detail_sheet.cell(row=i,column=14).value
        # d.penalty_runs = detail_sheet.cell(row=i,column=15).value
        # d.batsman_runs = detail_sheet.cell(row=i,column=16).value
        # d.extra_runs = detail_sheet.cell(row=i,column=17).value
        # d.total_runs = detail_sheet.cell(row=i,column=18).value
        # d.player_dismissed = detail_sheet.cell(row=i,column=19).value
        # d.dismissal_kind = detail_sheet.cell(row=i,column=20).value
        # d.fielder = detail_sheet.cell(row=i,column=21).value
        # d.match=Matches.objects.get(matchid=detail_sheet.cell(row=i,column=1).value)

        # d.save()
    db.commit()
    db.close()



if __name__=='__main__':
   main()




