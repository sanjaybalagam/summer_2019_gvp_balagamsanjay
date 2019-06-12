import MySQLdb
import click
import openpyxl


db= MySQLdb.connect(host='127.0.0.1', user="root", passwd='admin', port=3306)

cur = db.cursor()


@click.group()
def assign5():
    pass


@assign5.command()
@click.argument('dbname')
def createdb(dbname):
    try:
        cur.execute('drop database '+dbname+';')
    except:
        pass
    cur.execute('create database '+dbname+';')
    cur.execute('use '+dbname+';')
    cur.execute('create table students( name char(30) NOT NULL, college char(20) NOT NULL, primary key(name));')
    cur.execute('create table results (name char(30) NOT NULL ,total int(3) NOT NULL, foreign key(name) references students(name));')

@assign5.command()
@click.argument('dbname')
def dropdb(dbname):
    cur.execute('drop database '+dbname+';')


@assign5.command()
@click.argument('dbname')
def importtodb(dbname):
    i=1
    try:
        cur.execute('use '+dbname+';')
        cur1= db.cursor()
        cur1.execute('use '+dbname+";")
        wb1=openpyxl.load_workbook('students.xlsx')
        wb2=openpyxl.load_workbook('abc.xlsx')
        sheet1=wb1['Current']
        sheet2=wb2['Sheet']
        try:
            for row1 in range(2,sheet1.max_row+1):
                name=sheet1.cell(row= row1, column =4).value
                clg = sheet1.cell(row=row1, column=2).value
                cur.execute('insert into students values("'+name+'" , "'+clg+'" );')


                total=sheet2.cell(row=row1, column=6).value
                cur1.execute('insert into results values("'+name+'" , '+total+');')
        except:
            print('no record')
        cur.execute('commit;')
        cur1.execute('commit;')

    except Exception as a:
        print(dbname+" doesnt exist")


@assign5.command()
def collegestats(dbname):
    cur.execute('use '+dbname+';')
    cur.execute('')


if __name__ == '__main__':
    assign5()