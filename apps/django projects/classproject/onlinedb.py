import click
import django
import openpyxl
import os


os.environ.setdefault('DJANGO_SETTINGS_MODULE', "classproject.settings")
django.setup()

from onlineapp.models import *

colleges=[]
c= None
@click.command()
def onlinedb():
    wb1=openpyxl.load_workbook('assign5/students.xlsx')
    wb2=openpyxl.load_workbook('assign5/abc.xlsx')
    sheet1=wb1['Current']
    sheet2=wb2['Sheet']

    for row1 in range(2,sheet1.max_row+1):
        name=sheet1.cell(row= row1, column =4).value
        clg = sheet1.cell(row=row1, column=2).value
        emailid=sheet1.cell(row=row1, column=3).value
        stu_dbname=sheet1.cell(row=row1, column=4).value

        mark1=sheet2.cell(row=row1, column=2).value
        mark2 = sheet2.cell(row=row1, column=3).value
        mark3 = sheet2.cell(row=row1, column=4).value
        mark4 = sheet2.cell(row=row1, column=5).value
        total=sheet2.cell(row=row1, column=6).value

        if (clg not in colleges):
            c=College()
            c.name=clg
            c.location='default'
            c.acronym=clg
            c.contact=(clg+'@edu.com')
            c.save()
            colleges.append(clg)


        s= Student()
        s.name=name
        s.db_folder=stu_dbname
        s.dob='1998-01-01'
        s.email=emailid
        s.dropped_out=False
        s.college=c
        s.save()

        mk = MockTest1()
        mk.name=s
        mk.problem1= mark1
        mk.problem2 = mark2
        mk.problem3 = mark3
        mk.problem4 = mark4
        mk.total = total
        mk.student=s
        mk.save()



if __name__ == '__main__':
    onlinedb()




